#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Run one online vLLM benchmark case.

This script benchmarks an already running vLLM OpenAI-compatible server.

Typical server:
  vllm serve /path/to/model \
    --served-model-name Qwen3-30B-A3-w8a8 \
    --tensor-parallel-size 1 \
    --data-parallel-size 2 \
    --host 0.0.0.0 \
    --port 8000

Online metrics:
  - request throughput
  - output token throughput
  - total token throughput
  - end-to-end latency
  - TTFT, when streaming is enabled
  - TPOT
  - ITL, approximated from streaming chunks

Note:
  TTFT / ITL require --stream. Without streaming, the script can only report
  end-to-end latency and token throughput.
"""

import argparse
import csv
import json
import math
import os
import statistics
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import requests
from transformers import AutoTokenizer
import yaml


def parse_yaml_config(config_path: str | Path) -> Dict[str, Any]:
    """Parse and return YAML config."""
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    if not isinstance(config, dict):
        raise SystemExit(f"Config file must be a YAML object, got {type(config).__name__}")
    return config


def validate_client_config(config: Dict[str, Any]) -> List[str]:
    """Validate client config and return list of error messages."""
    errors = []

    if "base-url" not in config or not config["base-url"]:
        errors.append("client.base-url is required")
    if "model" not in config or not config["model"]:
        errors.append("client.model is required")

    return errors


def yaml_to_cli_args(client_config: Dict[str, Any]) -> List[str]:
    """Convert YAML client config dict to list of CLI argument strings."""
    args: List[str] = []

    for yaml_key, value in client_config.items():
        if value is None:
            continue

        # Skip internal fields that have no CLI equivalent
        if yaml_key in ("dataset-name", "save-result", "result-filename"):
            continue

        # Handle stream specially: True -> --stream, False -> --no-stream
        if yaml_key == "stream":
            if value is True:
                args.append("--stream")
            elif value is False:
                args.append("--no-stream")
            continue

        # All other YAML fields use -- prefix to become CLI args
        if isinstance(value, bool):
            if value:
                args.append(f"--{yaml_key}")
        else:
            args.append(f"--{yaml_key}")
            args.append(str(value))

    return args


def percentile(values: Sequence[float], p: float) -> Optional[float]:
    if not values:
        return None
    xs = sorted(values)
    if len(xs) == 1:
        return xs[0]
    k = (len(xs) - 1) * p / 100.0
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return xs[int(k)]
    return xs[f] * (c - k) + xs[c] * (k - f)


def mean(values: Sequence[float]) -> Optional[float]:
    return statistics.mean(values) if values else None


def median(values: Sequence[float]) -> Optional[float]:
    return statistics.median(values) if values else None


def ms(x: Optional[float]) -> Optional[float]:
    return None if x is None else x * 1000.0


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2, default=str)


def append_csv(path: Path, row: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = path.exists()
    with open(path, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row.keys()))
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

def normalize_excel_value(value: Any) -> Any:
    """Convert complex values to Excel-friendly scalar values."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def append_xlsx(path: Path, row: Dict[str, Any], sheet_name: str = "summary") -> None:
    """Append one summary row to an xlsx file.

    If the file does not exist, create it and write the header first.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    headers = list(row.keys())

    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)

    # 如果已有文件但表头为空，补表头
    if ws.max_row == 0:
        ws.append(headers)

    # 如果已有文件表头和当前 summary 字段不完全一致，按已有表头为准，
    # 新字段追加到表头末尾。
    existing_headers = [cell.value for cell in ws[1]]
    for h in headers:
        if h not in existing_headers:
            existing_headers.append(h)
            ws.cell(row=1, column=len(existing_headers), value=h)

    values = [normalize_excel_value(row.get(h)) for h in existing_headers]
    ws.append(values)

    # 简单美化
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # 自动列宽，限制最大宽度，避免路径太长撑爆
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    wb.save(path)

def load_prompt_file(path: str) -> List[str]:
    prompts: List[str] = []
    with open(path, "r", encoding="utf-8") as f:
        if path.endswith(".jsonl"):
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                if isinstance(obj, dict):
                    text = (
                        obj.get("prompt")
                        or obj.get("text")
                        or obj.get("content")
                        or obj.get("input")
                        or ""
                    )
                    if text:
                        prompts.append(str(text))
                elif isinstance(obj, str):
                    prompts.append(obj)
        else:
            text = f.read()
            # 用空行切分；如果没有空行，就整个文件作为一个 prompt
            parts = [x.strip() for x in text.split("\n\n") if x.strip()]
            prompts.extend(parts if parts else [text.strip()])
    return prompts


def make_synthetic_prompt(language: str) -> str:
    if language == "en":
        return (
            "This is a synthetic benchmark prompt for evaluating large language "
            "model inference performance. Please read the context carefully and "
            "answer the final question in a concise and structured way. "
        )
    return (
        "这是一个用于测试大语言模型推理性能的合成输入。请认真阅读上下文，"
        "并根据要求给出结构清晰、内容准确、表达简洁的回答。"
    )


def normalize_to_token_len(
    *,
    tokenizer: Any,
    text: str,
    target_len: int,
) -> Tuple[str, int]:
    """Repeat/truncate text to approximately exact target token length."""
    if target_len <= 0:
        ids = tokenizer.encode(text, add_special_tokens=False)
        return text, len(ids)

    ids = tokenizer.encode(text, add_special_tokens=False)
    if not ids:
        text = "hello"
        ids = tokenizer.encode(text, add_special_tokens=False)

    repeated: List[int] = []
    while len(repeated) < target_len:
        repeated.extend(ids)

    repeated = repeated[:target_len]
    normalized = tokenizer.decode(repeated, skip_special_tokens=True)
    actual_len = len(tokenizer.encode(normalized, add_special_tokens=False))
    return normalized, actual_len


def build_prompts(args: argparse.Namespace, tokenizer: Any) -> Tuple[List[str], List[int], List[str]]:
    if args.prompt_file:
        raw_prompts = load_prompt_file(args.prompt_file)
        if not raw_prompts:
            raise SystemExit(f"no prompts loaded from {args.prompt_file}")
    else:
        raw_prompts = [make_synthetic_prompt(args.synthetic_language)]

    prompts: List[str] = []
    input_tokens: List[int] = []
    sources: List[str] = []

    requested_n = args.num_prompts
    dataset_n = len(raw_prompts)

    if requested_n <= 0:
        n = dataset_n
    else:
        n = min(requested_n, dataset_n)

    if args.prompt_file and requested_n > dataset_n:
        print(
            f"[WARN] requested num_prompts={requested_n}, "
            f"but prompt_file only has {dataset_n} records. "
            f"Use real records only: num_prompts={n}"
        )
    args.num_prompts = n
    for i in range(n):
        src = raw_prompts[i]
        if args.normalize_file_prompts or not args.prompt_file:
            prompt, tok_len = normalize_to_token_len(
                tokenizer=tokenizer,
                text=src,
                target_len=args.input_len,
            )
        else:
            prompt = src
            tok_len = len(tokenizer.encode(prompt, add_special_tokens=False))

        prompts.append(prompt)
        input_tokens.append(tok_len)
        sources.append(args.prompt_file or "synthetic")

    return prompts, input_tokens, sources


def parse_sse_lines(response: requests.Response):
    """Yield parsed JSON objects from OpenAI-compatible SSE stream."""
    for raw_line in response.iter_lines(decode_unicode=True):
        if raw_line is None:
            continue
        line = raw_line.strip()
        if not line:
            continue
        if not line.startswith("data:"):
            continue

        data = line[len("data:"):].strip()
        if data == "[DONE]":
            break

        try:
            yield json.loads(data)
        except json.JSONDecodeError:
            continue


def extract_content_from_stream_chunk(obj: Dict[str, Any], endpoint: str) -> str:
    try:
        choice = obj.get("choices", [{}])[0]
        if endpoint == "chat":
            delta = choice.get("delta", {})
            return delta.get("content") or ""
        else:
            return choice.get("text") or ""
    except Exception:
        return ""


def extract_content_from_nonstream_response(obj: Dict[str, Any], endpoint: str) -> str:
    choice = obj.get("choices", [{}])[0]
    if endpoint == "chat":
        msg = choice.get("message", {})
        return msg.get("content") or ""
    return choice.get("text") or ""


def build_payload(
    *,
    args: argparse.Namespace,
    prompt: str,
    stream: bool,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "model": args.model,
        "temperature": args.temperature,
        "top_p": args.top_p,
        "max_tokens": args.output_len,
        "stream": stream,
    }

    if args.top_k is not None:
        # vLLM OpenAI server 通常支持 extra_body 风格参数，
        # 但直接传 top_k 在部分版本也可用；不支持时服务会报错。
        payload["top_k"] = args.top_k

    if args.seed is not None:
        payload["seed"] = args.seed

    if args.ignore_eos:
        payload["ignore_eos"] = True

    if args.endpoint == "chat":
        payload["messages"] = [
            {"role": "user", "content": prompt}
        ]
    else:
        payload["prompt"] = prompt

    # 有些 vLLM 版本支持流式 usage；不支持也通常会忽略或报错。
    if stream and args.include_usage:
        payload["stream_options"] = {"include_usage": True}

    return payload


def one_request(
    *,
    request_idx: int,
    prompt: str,
    input_tokens: int,
    prompt_source: str,
    args: argparse.Namespace,
    tokenizer: Any,
) -> Dict[str, Any]:
    url = args.base_url.rstrip("/") + ("/v1/chat/completions" if args.endpoint == "chat" else "/v1/completions")
    headers = {"Content-Type": "application/json"}
    if args.api_key:
        headers["Authorization"] = f"Bearer {args.api_key}"
    payload = build_payload(args=args, prompt=prompt, stream=args.stream)
    start = time.perf_counter()
    
    if args.stream:
        output_text, first_token_time, chunk_times, status_code, error = _handle_stream_request(
            url, headers, payload, args.timeout, args.endpoint
        )
    else:
        output_text, status_code, error = _handle_nonstream_request(
            url, headers, payload, args.timeout, args.endpoint
        )
        first_token_time = None
        chunk_times = []
    
    end = time.perf_counter()
    output_tokens = len(tokenizer.encode(output_text, add_special_tokens=False))
    
    return _build_result_dict(
        request_idx=request_idx,
        success=error is None,
        status_code=status_code,
        error=error,
        prompt_source=prompt_source,
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        start=start,
        end=end,
        first_token_time=first_token_time,
        chunk_times=chunk_times,
        preview_chars=args.preview_chars,
        print_output=args.print_output,
        output_text=output_text, 
    )


def _handle_stream_request(
    url: str,
    headers: Dict[str, str],
    payload: Dict[str, Any],
    timeout: float,
    endpoint: str,
) -> Tuple[str, Optional[float], List[float], Optional[int], Optional[str]]:
    output_text_parts: List[str] = []
    chunk_times: List[float] = []
    first_token_time: Optional[float] = None
    status_code: Optional[int] = None
    error: Optional[str] = None
    
    try:
        with requests.post(url, headers=headers, json=payload, timeout=timeout, stream=True) as resp:
            status_code = resp.status_code
            if resp.status_code != 200:
                error = resp.text[:1000]
            else:
                for obj in parse_sse_lines(resp):
                    now = time.perf_counter()
                    text = extract_content_from_stream_chunk(obj, endpoint)
                    if text:
                        output_text_parts.append(text)
                        chunk_times.append(now)
                        if first_token_time is None:
                            first_token_time = now
    except Exception as exc:
        error = repr(exc)
    
    return "".join(output_text_parts), first_token_time, chunk_times, status_code, error


def _handle_nonstream_request(
    url: str,
    headers: Dict[str, str],
    payload: Dict[str, Any],
    timeout: float,
    endpoint: str,
) -> Tuple[str, Optional[int], Optional[str]]:
    output_text = ""
    status_code: Optional[int] = None
    error: Optional[str] = None
    
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
        status_code = resp.status_code
        if resp.status_code != 200:
            error = resp.text[:1000]
        else:
            obj = resp.json()
            output_text = extract_content_from_nonstream_response(obj, endpoint)
    except Exception as exc:
        error = repr(exc)
    
    return output_text, status_code, error


def _build_result_dict(
    request_idx: int,
    success: bool,
    status_code: Optional[int],
    error: Optional[str],
    prompt_source: str,
    input_tokens: int,
    output_tokens: int,
    start: float,
    end: float,
    first_token_time: Optional[float],
    chunk_times: List[float],
    preview_chars: int,
    print_output: bool,
    output_text: str,
) -> Dict[str, Any]:
    latency_s = end - start
    ttft_s = None if first_token_time is None else first_token_time - start
    
    if first_token_time is not None and output_tokens > 1:
        tpot_s = (end - first_token_time) / (output_tokens - 1)  # 去掉 max，因为已经判断 > 1
    else:
        tpot_s = None
    
    itls: List[float] = []
    if len(chunk_times) >= 2:
        itls = [chunk_times[i] - chunk_times[i - 1] for i in range(1, len(chunk_times))]
    
    return {
        "request_idx": request_idx,
        "success": success,
        "status_code": status_code,
        "error": error,
        "prompt_source": prompt_source,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
        "latency_s": latency_s,
        "ttft_s": ttft_s,
        "tpot_s": tpot_s,
        "itl_avg_s": mean(itls),
        "itl_p50_s": median(itls),
        "itl_p90_s": percentile(itls, 90),
        "itl_p99_s": percentile(itls, 99),
        "num_stream_chunks": len(chunk_times),
        "output_preview": output_text[:preview_chars] if print_output else "",
    }

def summarize(
    *,
    args: argparse.Namespace,
    run_id: str,
    case_id: str,
    measured_wall_time_s: float,
    records: List[Dict[str, Any]],
) -> Dict[str, Any]:
    ok = [r for r in records if r["success"]]
    fail = [r for r in records if not r["success"]]

    latencies = [r["latency_s"] for r in ok]
    ttfts = [r["ttft_s"] for r in ok if r.get("ttft_s") is not None]
    tpots = [r["tpot_s"] for r in ok if r.get("tpot_s") is not None]
    itl_avgs = [r["itl_avg_s"] for r in ok if r.get("itl_avg_s") is not None]

    total_input_tokens = sum(r["input_tokens"] for r in ok)
    total_output_tokens = sum(r["output_tokens"] for r in ok)
    total_tokens = total_input_tokens + total_output_tokens

    success_count = len(ok)

    summary: Dict[str, Any] = {
        "run_id": run_id,
        "case_id": case_id,
        "bench_mode": "online",
        "endpoint": args.endpoint,
        "base_url": args.base_url,
        "model": args.model,
        "tokenizer": args.tokenizer or args.model,
        "num_prompts": args.num_prompts,
        "max_concurrency": args.max_concurrency,
        "request_rate": args.request_rate,
        "input_len_target": args.input_len,
        "output_len_target": args.output_len,
        "temperature": args.temperature,
        "top_p": args.top_p,
        "ignore_eos": args.ignore_eos,
        "stream": args.stream,
        "hardware": args.hardware,
        "tester": args.tester,

        "success_count": success_count,
        "failed_count": len(fail),
        "measured_wall_time_s": measured_wall_time_s,

        "total_input_tokens": total_input_tokens,
        "total_output_tokens": total_output_tokens,
        "total_tokens": total_tokens,

        "request_throughput_req_s": success_count / measured_wall_time_s if measured_wall_time_s > 0 else None,
        "output_tokens_per_s": total_output_tokens / measured_wall_time_s if measured_wall_time_s > 0 else None,
        "total_tokens_per_s": total_tokens / measured_wall_time_s if measured_wall_time_s > 0 else None,

        "latency_avg_ms": ms(mean(latencies)),
        "latency_p50_ms": ms(percentile(latencies, 50)),
        "latency_p90_ms": ms(percentile(latencies, 90)),
        "latency_p99_ms": ms(percentile(latencies, 99)),

        "ttft_avg_ms": ms(mean(ttfts)),
        "ttft_p50_ms": ms(percentile(ttfts, 50)),
        "ttft_p90_ms": ms(percentile(ttfts, 90)),
        "ttft_p99_ms": ms(percentile(ttfts, 99)),

        "tpot_avg_ms": ms(mean(tpots)),
        "tpot_p50_ms": ms(percentile(tpots, 50)),
        "tpot_p90_ms": ms(percentile(tpots, 90)),
        "tpot_p99_ms": ms(percentile(tpots, 99)),

        "itl_avg_ms": ms(mean(itl_avgs)),
        "itl_p50_ms": ms(percentile(itl_avgs, 50)),
        "itl_p90_ms": ms(percentile(itl_avgs, 90)),
        "itl_p99_ms": ms(percentile(itl_avgs, 99)),
    }

    return summary


def health_check(args: argparse.Namespace) -> None:
    url = args.base_url.rstrip("/") + "/health"
    try:
        resp = requests.get(url, timeout=10)
        print(f"[INFO] health: status={resp.status_code}")
    except Exception as exc:
        print(f"[WARN] health check failed: {exc!r}")


def list_models(args: argparse.Namespace) -> None:
    url = args.base_url.rstrip("/") + "/v1/models"
    try:
        resp = requests.get(url, timeout=10)
        print(f"[INFO] models status={resp.status_code}")
        if resp.status_code == 200:
            obj = resp.json()
            ids = [x.get("id") for x in obj.get("data", [])]
            print(f"[INFO] served models: {ids}")
    except Exception as exc:
        print(f"[WARN] list models failed: {exc!r}")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Online vLLM benchmark via OpenAI-compatible HTTP API")

    p.add_argument("--config", default="", help="YAML config file path")
    p.add_argument("--dry-run", action="store_true", help="validate config and exit without running")

    p.add_argument("--base-url", default="http://127.0.0.1:8000")
    p.add_argument("--model", default="", help="served model name, e.g. Qwen3-30B-A3B-w8a8")
    p.add_argument("--tokenizer", default="", help="tokenizer path/name, default: --model")
    p.add_argument("--api-key", default=os.getenv("OPENAI_API_KEY", ""))

    p.add_argument("--endpoint", choices=["chat", "completion"], default="chat")
    p.add_argument("--stream", action="store_true", default=True)
    p.add_argument("--no-stream", dest="stream", action="store_false")
    p.add_argument("--include-usage", action="store_true")

    p.add_argument("--num-prompts", type=int, default=16)
    p.add_argument("--max-concurrency", type=int, default=1)
    p.add_argument("--request-rate", type=float, default=0.0, help="0 means submit as fast as possible")

    p.add_argument("--input-len", type=int, default=512)
    p.add_argument("--output-len", type=int, default=256)
    p.add_argument("--prompt-file", default="")
    p.add_argument("--normalize-file-prompts", action="store_true")
    p.add_argument("--synthetic-language", choices=["zh", "en"], default="zh")

    p.add_argument("--temperature", type=float, default=0.0)
    p.add_argument("--top-p", type=float, default=1.0)
    p.add_argument("--top-k", type=int, default=None)
    p.add_argument("--seed", type=int, default=None)
    p.add_argument("--ignore-eos", action="store_true")
    p.add_argument("--timeout", type=float, default=600.0)

    p.add_argument("--run-id", default="")
    p.add_argument("--case-id", default="")
    p.add_argument("--tester", default=os.getenv("USER", "unknown"))
    p.add_argument("--hardware", default="Ascend 910B")
    p.add_argument("--output-dir", default="bench_results_online")
    p.add_argument("--print-output", action="store_true")
    p.add_argument("--preview-chars", type=int, default=300)

    p.add_argument("--skip-health-check", action="store_true")
    p.add_argument("--skip-list-models", action="store_true")

    args = p.parse_args(argv)

    # Auto-generate run_id and case_id if not provided
    if not args.run_id:
        args.run_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]

    if not args.case_id:
        args.case_id = (
            f"online_{args.endpoint}"
            f"_in{args.input_len}_out{args.output_len}"
            f"_n{args.num_prompts}_c{args.max_concurrency}"
        )

    return args


def run_single_benchmark(client_args: argparse.Namespace) -> None:
    """Run a single benchmark with the given args."""
    output_dir = Path(client_args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print("========== vLLM online benchmark ==========")
    print(f"base_url: {client_args.base_url}")
    print(f"model: {client_args.model}")
    print(f"endpoint: {client_args.endpoint}")
    print(f"stream: {client_args.stream}")
    print(f"num_prompts: {client_args.num_prompts}")
    print(f"max_concurrency: {client_args.max_concurrency}")
    print(f"request_rate: {client_args.request_rate}")
    print(f"input_len: {client_args.input_len}")
    print(f"output_len: {client_args.output_len}")

    if not client_args.skip_health_check:
        health_check(client_args)

    if not client_args.skip_list_models:
        list_models(client_args)

    tokenizer_path = client_args.tokenizer or client_args.model
    print(f"[INFO] loading tokenizer: {tokenizer_path}")
    tokenizer = AutoTokenizer.from_pretrained(tokenizer_path, trust_remote_code=True)

    print("[INFO] building prompts ...")
    prompts, input_tokens, prompt_sources = build_prompts(client_args, tokenizer)

    records: List[Dict[str, Any]] = []

    print("[INFO] running online requests ...")
    measured_start = time.perf_counter()

    with ThreadPoolExecutor(max_workers=client_args.max_concurrency) as executor:
        futures = []

        for i, prompt in enumerate(prompts):
            fut = executor.submit(
                one_request,
                request_idx=i,
                prompt=prompt,
                input_tokens=input_tokens[i],
                prompt_source=prompt_sources[i],
                args=client_args,
                tokenizer=tokenizer,
            )
            futures.append(fut)

            if client_args.request_rate and client_args.request_rate > 0:
                time.sleep(1.0 / client_args.request_rate)

        for fut in as_completed(futures):
            rec = fut.result()
            records.append(rec)

            status = "OK" if rec["success"] else "FAIL"
            print(
                f"[{status}] idx={rec['request_idx']} "
                f"latency={rec['latency_s']:.3f}s "
                f"ttft={rec['ttft_s']} "
                f"in={rec['input_tokens']} "
                f"out={rec['output_tokens']} "
                f"err={rec['error'][:120] if rec['error'] else ''}"
            )

    measured_wall_time_s = time.perf_counter() - measured_start

    # 保持 request_idx 顺序
    records = sorted(records, key=lambda x: x["request_idx"])

    summary = summarize(
        args=client_args,
        run_id=client_args.run_id,
        case_id=client_args.case_id,
        measured_wall_time_s=measured_wall_time_s,
        records=records,
    )

    stem = f"{client_args.case_id}_{client_args.run_id}"
    summary_path = output_dir / f"{stem}.summary.json"
    requests_path = output_dir / f"{stem}.requests.jsonl"
    csv_path = output_dir / "summary.csv"
    xlsx_path = output_dir / "summary.xlsx"

    write_json(summary_path, {"summary": summary})
    with open(requests_path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    append_csv(csv_path, summary)
    append_xlsx(xlsx_path, summary)

    print("\n========== SUMMARY ==========")
    for key in [
        "run_id",
        "case_id",
        "bench_mode",
        "endpoint",
        "model",
        "num_prompts",
        "max_concurrency",
        "request_rate",
        "input_len_target",
        "output_len_target",
        "success_count",
        "failed_count",
        "measured_wall_time_s",
        "total_input_tokens",
        "total_output_tokens",
        "request_throughput_req_s",
        "output_tokens_per_s",
        "total_tokens_per_s",
        "latency_avg_ms",
        "latency_p50_ms",
        "latency_p90_ms",
        "latency_p99_ms",
        "ttft_avg_ms",
        "ttft_p50_ms",
        "ttft_p90_ms",
        "ttft_p99_ms",
        "tpot_avg_ms",
        "tpot_p50_ms",
        "tpot_p90_ms",
        "tpot_p99_ms",
        "itl_avg_ms",
        "itl_p50_ms",
        "itl_p90_ms",
        "itl_p99_ms",
    ]:
        print(f"{key}: {summary.get(key)}")

    print("\n[INFO] outputs:")
    print(f"summary : {summary_path}")
    print(f"requests: {requests_path}")
    print(f"csv     : {csv_path}")
    print(f"xlsx    : {xlsx_path}")


def load_yaml_clients(config_path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Load and validate client configs from YAML file.
    
    Returns:
        Tuple of (clients_list, global_config)
    """
    try:
        config = parse_yaml_config(config_path)
    except FileNotFoundError:
        raise SystemExit(f"Config file not found: {config_path}")
    except yaml.YAMLError as exc:
        raise SystemExit(f"Failed to parse YAML: {exc}")

    if "client" not in config or not config["client"]:
        raise SystemExit("Missing required section: client")

    clients = config["client"]
    if isinstance(clients, dict):
        clients = [clients]
    elif not isinstance(clients, list):
        raise SystemExit(f"client must be a dict or list of dicts, got {type(clients).__name__}")

    # Validate all client configs
    for i, client in enumerate(clients):
        errors = validate_client_config(client)
        if errors:
            print(f"Client[{i}] validation failed:")
            for err in errors:
                print(f"  - {err}")
            raise SystemExit(1)

    return clients, config


def main() -> None:
    import sys

    # Pre-check for --config in command line (before full parse_args)
    config_path = ""
    for i, arg in enumerate(sys.argv):
        if arg == "--config" and i + 1 < len(sys.argv):
            config_path = sys.argv[i + 1]
            break

    # If --config provided, read YAML and run benchmarks
    if config_path:
        clients, _ = load_yaml_clients(config_path)

        # Check for dry-run
        if "--dry-run" in sys.argv:
            print("========== DRY RUN: Generated CLI Commands ==========")
            for i, client_config in enumerate(clients):
                cli_args = yaml_to_cli_args(client_config)
                full_cmd = f"python run_online_bench.py --config {config_path} " + " ".join(cli_args)
                print(f"\n--- Client {i+1}/{len(clients)} ---")
                print(full_cmd)
            print("\n[DRY RUN] No requests will be sent.")
            raise SystemExit(0)

        # YAML config mode: iterate over all clients
        for i, client_config in enumerate(clients):
            print(f"\n{'=' * 60}")
            print(f"Running client {i + 1}/{len(clients)}")
            print(f"{'=' * 60}")

            # Convert YAML client config to CLI args
            cli_args = yaml_to_cli_args(client_config)
            # Prepend project default arguments, then YAML args override them
            default_args = ["--config", config_path]
            client_args = parse_args(default_args + cli_args)

            # Auto-generate run_id and case_id
            if not client_args.run_id:
                client_args.run_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
            if not client_args.case_id:
                client_args.case_id = (
                    f"online_{client_args.endpoint}"
                    f"_in{client_args.input_len}_out{client_args.output_len}"
                    f"_n{client_args.num_prompts}_c{client_args.max_concurrency}"
                )

            run_single_benchmark(client_args)
    else:
        # CLI mode: use original args directly
        args = parse_args()
        run_single_benchmark(args)


if __name__ == "__main__":
    main()